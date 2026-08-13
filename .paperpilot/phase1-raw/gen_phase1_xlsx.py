#!/usr/bin/env python3
"""Phase 1 issue 分类清单 → xlsx。含 126 issue 的分类 + 依据 + dev-reviewer 交叉结果。"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.abspath(__file__))
PAPER = os.path.dirname(os.path.dirname(ROOT))
MANIFEST = json.load(open(os.path.join(ROOT, 'manifest.json'), encoding='utf-8'))
CASES = {(x['vendor'], x['num']): x for x in json.load(open(os.path.join(PAPER, '.paperpilot', 'phase2-rerun', 'cases_index.json'), encoding='utf-8'))}

GROUP_A = {'TP_FIXED_PR'}
GROUP_B = {'TP_ACK_OPEN', 'TP_ACK_CLOSED_NOFIX', 'TP_DUP_TRACKED'}
GROUP_C = {'FP_BY_DESIGN', 'FP_NOT_REPRO', 'BY_DESIGN'}
EXCLUDED = {'SELF_PR_CLOSED', 'SELF_PR_OPEN'}

BASIS = {
    'TP_FIXED_PR': 'A 真 bug：已被合并 PR 修复',
    'TP_ACK_OPEN': 'B 真 bug：维护者确认但未修(open)',
    'TP_ACK_CLOSED_NOFIX': 'B 真 bug：维护者确认但未修(closed nofix)',
    'TP_DUP_TRACKED': 'B 真 bug：重复，已在别处追踪',
    'FP_BY_DESIGN': 'C 误报：维护者判定 by-design',
    'FP_NOT_REPRO': 'C 误报：不可复现',
    'BY_DESIGN': 'C 误报：by-design(无 bug)',
    'PENDING_SELF_LABELED': 'D 未裁决：报告者自标，待维护者',
    'SELF_CLOSED': 'D 未裁决：报告者自行关闭',
    'STALE_NO_FIX': 'D 未裁决：过期未修',
    'OPEN_NO_LABEL': 'D 未裁决：open 无标签',
    'SELF_PR_CLOSED': '排除：自提 PR(非 issue)',
    'SELF_PR_OPEN': '排除：自提 PR(非 issue)',
}


def group_of(cat):
    if cat in EXCLUDED:
        return 'EXCLUDED'
    if cat in GROUP_A:
        return 'A'
    if cat in GROUP_B:
        return 'B'
    if cat in GROUP_C:
        return 'C'
    return 'D'


def gtlabel_of(cat):
    g = group_of(cat)
    return {'A': 'CONFIRMED', 'B': 'CONFIRMED', 'C': 'FALSE_POSITIVE'}.get(g, 'unadjudicated')


# dev-reviewer 交叉结果(只 scored 子集)
RUN = os.path.join(PAPER, '.paperpilot', 'phase2-rerun', 'run')
def dev_verdict(vendor, num):
    e = CASES.get((vendor, num))
    if not e:
        return '', ''
    dr = os.path.join(RUN, 'results', vendor, e['version'], str(num), 'debate_logs', 'dev_review.json')
    if os.path.isfile(dr):
        try:
            d = json.load(open(dr, encoding='utf-8'))
            v = d['verdicts'][0]
            return v.get('verdict', ''), v.get('confidence', '')
        except Exception:
            pass
    return '', ''


wb = Workbook()
ws = wb.active
ws.title = 'issues'
headers = ['vendor', 'number', 'repo', 'title', 'state', 'state_reason', 'is_pr', 'created_at',
           'reported_version', 'version_source', 'gt_category', 'group', 'gt_label',
           'classification_basis', 'dev_reviewer_verdict', 'dev_reviewer_confidence']
ws.append(headers)
for c in ws[1]:
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='D9E1F2')

for it in sorted(MANIFEST, key=lambda x: (x['repo'], x['number'])):
    vendor = it['repo'].split('/')[1]
    num = it['number']
    cat = it['gt_category']
    dv, dc = dev_verdict(vendor, num)
    ws.append([vendor, num, it['repo'], it.get('title', ''), it.get('state', ''), it.get('state_reason', ''),
               it.get('is_pr', ''), it.get('created_at', ''), it.get('reported_version', ''),
               it.get('version_source') or '', cat, group_of(cat), gtlabel_of(cat),
               BASIS.get(cat, cat), dv, dc])
ws.column_dimensions['D'].width = 50
ws.column_dimensions['N'].width = 36
ws.freeze_panes = 'A2'

# legend sheet
lg = wb.create_sheet('legend')
lg.append(['group', 'gt_label', 'gt_category', '含义/依据', '计数'])
for c in lg[1]:
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', fgColor='D9E1F2')
import collections
cnt = collections.Counter(it['gt_category'] for it in MANIFEST)
order = [('A', 'CONFIRMED', 'TP_FIXED_PR'), ('B', 'CONFIRMED', 'TP_ACK_OPEN'),
         ('B', 'CONFIRMED', 'TP_ACK_CLOSED_NOFIX'), ('B', 'CONFIRMED', 'TP_DUP_TRACKED'),
         ('C', 'FALSE_POSITIVE', 'FP_BY_DESIGN'), ('C', 'FALSE_POSITIVE', 'FP_NOT_REPRO'),
         ('C', 'FALSE_POSITIVE', 'BY_DESIGN'), ('D', 'unadjudicated', 'PENDING_SELF_LABELED'),
         ('D', 'unadjudicated', 'SELF_CLOSED'), ('D', 'unadjudicated', 'STALE_NO_FIX'),
         ('D', 'unadjudicated', 'OPEN_NO_LABEL'), ('EXCLUDED', '—', 'SELF_PR_CLOSED'),
         ('EXCLUDED', '—', 'SELF_PR_OPEN')]
for g, lab, cat in order:
    lg.append([g, lab, cat, BASIS.get(cat, cat), cnt.get(cat, 0)])
lg.append([])
lg.append(['合计 A∪B∪C (scored 用于指标)', '', '', '', sum(cnt[c] for c in GROUP_A | GROUP_B | GROUP_C)])
lg.append(['合计 D (unadjudicated)', '', '', '', sum(cnt[c] for c in ['PENDING_SELF_LABELED', 'SELF_CLOSED', 'STALE_NO_FIX', 'OPEN_NO_LABEL'])])
lg.column_dimensions['D'].width = 40

out = os.path.join(PAPER, 'data', 'phase1_issue_classification.xlsx')
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print('wrote', out)
print('rows:', len(MANIFEST), '| A∪B∪C:', sum(1 for it in MANIFEST if group_of(it['gt_category']) in 'ABC'))
